import requests
from bs4 import BeautifulSoup as bp
import json
import openpyxl
import math


def main():  #主執行程式
    session = getSession() #取得requests
    stockAry = getstockID() #取的txt檔案中的 所需要的股票ID
    obj = getstockInfo(session, stockAry)

    setUpExcel(obj)


def getSession(): 
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0'
    }
    session = requests.session()
    session.headers.update(header)
    return session


def getstockInfo(session, stockAry):
    obj = {}
    for stock in stockAry:
        url = f'https://tw.stock.yahoo.com/quote/{stock}.TW' #取得股票名字、當前價格

        res = session.get(url)
        soup = bp(res.text, 'html.parser')

        title = soup.select('#main-0-QuoteHeader-Proxy')[0].find('h1').get_text()
        point = soup.select('#main-0-QuoteHeader-Proxy')[0].select('.Fz\(32px\)')[0].get_text()
        
        #取得股票配息、歷史配息
        interestUrl = f'https://tw.stock.yahoo.com/_td-stock/api/resource/ApacFinanceServices.etfInfo;etfRegion=tw;etfSymbol={stock}.TW?bkt=&device=desktop&ecma=modern&feature=enableGAMAds%2CenableGAMEdgeToEdge%2CenableEvPlayer&intl=tw&lang=zh-Hant-TW&partner=none'
        
        resJSON = session.get(interestUrl).text
        jsonTurnObj = json.loads(resJSON)

        dividend = float(jsonTurnObj["dividend"]["last"])
        historyDividend = jsonTurnObj["dividend"]["historical"]

        obj[stock] = [title, float(point), dividend, [historyDividend]]  # { 股票編號: [ 股票名字, 當前價格, 當前or前一次配息, [ 歷史配息 ] ]}

    return obj

def getstockID():  
    stockAry = []
    with open('stockID.txt', encoding='utf-8') as id:
        st = id.readlines()[0].split(',')
        for s in st:
            stockAry.append(s)
    return stockAry


def setUpExcel(obj):
    wb = openpyxl.Workbook()
    ws = wb.active

    column = ['股票編號', '股票名稱', '股票價格', '配息', '試算存10000', '股數', '殖利率']
    ws.append(column)

    count = 2
    for key, value in obj.items():
        ws.cell(count, 1).value = key
        ws.cell(count, 2).value = value[0]
        ws.cell(count, 3).value = value[1]
        ws.cell(count, 4).value = value[2]
        
        stock = math.floor(10000 / value[1])
        ws.cell(count, 6).value = stock
        ws.cell(count, 7).value = format((value[2] * stock) / value[1], '.2f')
        
        count += 1

    wb.save('123.xlsx')



if __name__ == '__main__':
    main()